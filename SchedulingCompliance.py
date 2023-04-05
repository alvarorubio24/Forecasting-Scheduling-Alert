import datetime as dt
from datetime import timedelta
from datetime import date
import datetime
import os
import time
from tkinter import *
from typing import final
import requests  # gives access to get/post methods
import win32com.client as client  # gives access to outlook
import openpyxl
import csv
from tabulate import tabulate
import sys
from tkinter import messagebox



import pandas as pd
import win32timezone
from pandas import ExcelWriter
import pprint
import pyodbc
import openpyxl







from SWATmetrics.SWATmetricsmain import SWATmain
from SAPlanningModule.SAPlanningmain import SAPlanningAlertAutomated
from RosteringTrackerModule.RosteringTrackermain import AutomatedRosteringTracker


#Forecasting&SchedulingAlert_1.8
#Owner: quiroalv
#Please note this coding is built by a beginner :)
# This coding is divided in 3 parts 
    # General Part
    # Scheduling Broadcaster 
    # Rostering Tracker
    # SA Planning



###########################
#######General Part########
###########################

#creating a directory for logs file
currentdirectory = os.getcwd()  #directory where the tool is located
user_login = os.getenv("username")
today = datetime.datetime.now().strftime("%Y-%m-%d") 
try:
    os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance')
    if not os.path.exists(today + " Compliance Uploads logs"):
        os.mkdir(today + " Compliance Uploads logs")

except Exception as e:
    messagebox.showwarning("Error", "Folder \Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance in shared drive could not be found\n\n"+ str(e))
    quit()


#this is for the scheduler who is running the app to know how much time is left for the module to be ran again
def countdown(t):
    print("\n")
    try:
        while t:
            mins, secs = divmod(t, 60)
            timer = '{:02d}:{:02d}'.format(mins, secs)
            print("Next scan will start in: " + timer, end="\r")
            time.sleep(1)
            t -= 1
    except Exception as e:
        messagebox.showwarning("Error", "Countdown function is not working, contact BIA team\n\n"+ str(e))
        quit()

      
  

#############################
#Scheduling Broadcaster Part#
#############################

def SchedulingBroadcaster():
    #these variables will be shared with Rostering Tracker part
    global list_of_emails   
    global dateOFD
    global DS2nduploadlist
    global DS2nduploadlistUK
    global DS3rduploadlist


    #this is for the getting the secs for the countdown function
    global Searchsecs

    try:
        print("----- Extra task owner: " + user_login + " -----")
        print("\nInitializing Scheduling Broadcaster... \n")
        os.chdir(currentdirectory)
        namefile = "data_config.xlsx" #there must be a file called data_config in the same folder where tool is located
        wb = openpyxl.load_workbook(namefile, data_only=True) #load workbook and telling Python to read only
        wconfig = wb.worksheets[0]
        
        #getting the folders name from the config_file as well as email address and search secs
        foldername = wconfig.cell(row=2, column=2).value
        Searchsecs = wconfig.cell(row=10, column=2).value
        Emailending = wconfig.cell(row=11, column=2).value
        username = user_login + "@amazon" + Emailending

    except Exception as e:
        messagebox.showwarning("Error", "The file data_config.xlsx could not be found \nMake sure this file is under the same folder this tool is stored\n\n"+ str(e.__class__))
        quit()

    try:
        #getting the messages that will be broadcasted via chime room to the DS
        #these are stored in the shared folder in an excel file
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance')
        wMsgs = openpyxl.load_workbook("Scheduling_Messages.xlsx", data_only=True) #load workbook and telling Python to read only
        wMessages = wMsgs.worksheets[0]
        #getting the messages from excel file
        message1stupload = wMessages.cell(row=2, column=1).value
        message2ndupload = wMessages.cell(row=2, column=3).value
        message2nduploadUK = wMessages.cell(row=2, column=5).value
        message3rdupload = wMessages.cell(row=2, column=7).value

    except Exception as e:
        messagebox.showwarning("Error", "File Scheduling_Messages.xlsx couldnt be found in \Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance in shared drive\n\n"+ str(e.__class__))
        quit()
    
    
    try:
        #getting the countries that will be sent to, for each message
        DS1stuploadlist = []
        range1 = wMessages.iter_cols()
        count1 = 6
        #we want a list with all the possible subjects that we will receive regarding scheduling uploads. Here is how is done
        for row in range1:
            for cell in row:
                if cell.value is not None:
                    DS1stupload = wMessages.cell(row=count1, column=1).value #this is the country 
                    DS1stuploadlist.append(DS1stupload + " - Next Day - Daily 1st run ") #this is country + part of the subject of the email
                    count1 +=1
    except:
        pass

    DS2nduploadlist = []
    range1 = wMessages.iter_cols()
    count2 = 6
    try: 
        for row in range1:
            for cell in row:
                if cell.value is not None:
                    DS2ndupload = wMessages.cell(row=count2, column=3).value
                    DS2nduploadlist.append(DS2ndupload + " - Next Day - Daily 2nd run ") #this is part of the subject of the email
                    count2 +=1
    except:
        pass

    DS2nduploadlistUK = []
    range1 = wMessages.iter_cols()
    count3 = 6
    try:
        for row in range1:
            for cell in row:
                if cell.value is not None:
                    DS2nduploadUK = wMessages.cell(row=count3, column=5).value
                    DS2nduploadlistUK.append(DS2nduploadUK + " - Next Day - Daily 2nd run ") #this is part of the subject of the email
                    count3 +=1
    except:
        pass

    DS3rduploadlist = []
    range1 = wMessages.iter_cols()
    count4 = 6
    try:
        for row in range1:
            for cell in row:
                if cell.value is not None:
                    DS3rdupload = wMessages.cell(row=count4, column=7).value
                    DS3rduploadlist.append(DS3rdupload + " - Next Day - Daily 3rd run ") #this is part of the subject of the email
                    count4 +=1
    except:
        pass

    #sending webhooks    
    def webhook(country,messageupload):
        #getting the countries to search DS in the webhook addresses excel
        countrylist = []
        if country == "UK":
            countrylist.extend(("UK","IE"))
        if country == "DE":
            countrylist.extend(("DE","AT","NL"))
        if country == "FR":
            countrylist.extend(("FR","BE"))
        if country == "IT":
            countrylist.append("IT")
        if country == "ES":
            countrylist.append("ES")

        
        countrystr = ', '.join(map(str, countrylist))
        print("\nSending webhooks to " + str(countrystr) + " ...")
        try:
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\Chime Broadcaster')#webhook addresses
            webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
            webhookadresseslist = []
            webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
        except Exception as e:
            messagebox.showwarning("Error", "File Webhooks with countries.csv couldnt be found in \\Business Analyses\CentralOPS\PM Shift\DHP1\Chime Broadcaster in shared drive\n\n"+ str(e.__class__))
            quit()
        try:
            for row in webhookcsv:
                if (countrylist[0] == row[2]) and "AMXL" != row[3]:
                    webhookurlrow =  row
                    webhookurl = webhookurlrow[1]
                    webhookadresseslist.append(webhookurl)
        except:
            pass
        
        try:
            webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
            for row in webhookcsv:
                if (countrylist[1] == row[2]) and "AMXL" != row[3]:
                    webhookurlrow =  row
                    webhookurl = webhookurlrow[1]
                    webhookadresseslist.append(webhookurl)
        except:
            pass
        try:
            webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
            for row in webhookcsv:
                if (countrylist[2] == row[2]) and "AMXL" != row[3]:
                    webhookurlrow =  row
                    webhookurl = webhookurlrow[1]
                    webhookadresseslist.append(webhookurl)
        except:
            pass
        try:
            webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
            for row in webhookcsv:
                if (countrylist[3] == row[2]) and "AMXL" != row[3]:
                    webhookurlrow =  row
                    webhookurl = webhookurlrow[1]
                    webhookadresseslist.append(webhookurl)
        except:
            pass
        
        #these are the messages to broadcast
        if messageupload in DS1stuploadlist:
            data = {"Content": message1stupload}
        if messageupload in DS2nduploadlist:
            data = {"Content": message2ndupload}
        if messageupload in DS2nduploadlistUK:
            data = {"Content": message2nduploadUK}
        if messageupload in DS3rduploadlist:
            data = {"Content": message3rdupload}
        
        #and then we send the message to all the chime rooms involved
        for webhook in webhookadresseslist:
            try:
                result = False
                session = requests.session()
                params = {'format': 'application/json'}
                response = session.post(webhook, params=params, json=data)
                if response.status_code == 200:
                    result = True
            except:
                webhookcsv = csv.reader(open("Webhooks with countries.csv","r"))
                for row in webhookcsv:
                    if (webhook == row[1]):
                        webhookurlrow =  row
                        webhookurlDS = webhookurlrow[0]
                        print("\nWebhook could not be sent to "+ webhookurlDS)
        
        
        print("\nScheduling Upload Webhooks has been sent to "+ str(countrystr))

           

        
    #creating a logs file in the shared drive to store information
    try:
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
        f = open(today + " Scheduling Broadcaster logs.csv","a", newline="") 
        writer = csv.writer(f)
        f.close()
        print("\nSearching for scheduling uploads...\n")
    except Exception as e:
        messagebox.showwarning("Error", "File Scheduling Broadcaster logs.csv couldnt be created under Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\" + today + " Compliance Uploads logs \nMake sure such directory in share drive exists \n\n"+ str(e.__class__))
        quit()
    
    
    #Reading outlook messages
    try:
        outlook = client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
    except Exception as e:
        messagebox.showwarning("Error", "Outlook couldnt be opened\nEnsure your Outlook is opened while running the tool\n\n"+ str(e.__class__))
        quit()

    try:
        rootFolder = namespace.Folders[username]
        
        emailFolder = rootFolder.Folders[foldername]
        
        messages = emailFolder.Items
        messages.Sort("[ReceivedTime]", True)

        list_of_emails = []
        date_of_emails = []
        
        #creating a list with all the subjects of the emails received
        #creating a list with all the dates from the subjects of the emails received (it will be used for the rostering tracker part)
        todayemails = datetime.date.today()
        for message in messages:
            if message.ReceivedTime.date() == todayemails: #emails received today only
                list_of_emails.append(message.Subject[0:30])
                date_of_emails.append(message.Subject[30:40])
    
    except Exception as e:
        messagebox.showwarning("Folders in Outlook Are Incorrectly Set Up", "Folders must be under your email address, not udner Inbox\nEnsure the name of the folders are exactly the same as the ones in data_config.xlsx\nAny space or typo can impact in the tool performance\n\n"+ str(e.__class__))
        quit()

    try:
        dateOFD = date_of_emails[0]
        dateOFD = datetime.datetime.strptime(dateOFD,"%Y-%m-%d")
    except:
        dateOFD = " " #if the OFD cannot be retrieved that means that emails received can be disregard anyway (it might be weekly or a reupload one)

    
    
    
    #get all values from csv
    #this csv will be updated adding new lines when a scheduling email is received and broadcasted
    #it will help us to know if the email was already received from previous runs
    try:
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
        excellist = []
        logsfile = csv.reader(open(today + " Scheduling Broadcaster logs.csv","r"))
        for row in logsfile:
            excellist.append(row[0]) #getting the first column
    
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
        excellisttodisplay = []
        logsfile = csv.reader(open(today + " Scheduling Broadcaster logs.csv","r"))
        for row in logsfile:
            excellisttodisplay.append(row)
    
    except:
        messagebox.showwarning("Error", "File Scheduling Broadcaster logs.csv couldnt be found under Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\" + today + " Compliance Uploads logs \nDelete all Rostering logs in such directory and restart tool \n\n"+ str(e.__class__))
        quit()  

    
    
    
    try:
        #displaying info from .csv in a nice table
        print(tabulate(excellisttodisplay,headers=['Email subject','Time Sent'], tablefmt='orgtbl'))        
        
        #merging all lists of the potential email subjects that we can receive regarding scheduling uploads
        uploadsubjectslist = []
        uploadsubjectslist = DS1stuploadlist + DS2nduploadlist + DS2nduploadlistUK + DS3rduploadlist
        
        #seeing the common values between the emails subject just received and all the potential ones that can be received (that are stored in excel file)
        list_of_emails_set = set(list(dict.fromkeys(list_of_emails)))
        intersection = list_of_emails_set.intersection(uploadsubjectslist)
        intersection_list = list(intersection) #the common values are in this list

    except Exception as e:
        messagebox.showwarning("Error", "There is a variable which is not defined\nSeek BIA team assistance\n\n"+ str(e))
        quit()
    
    
    timenow = datetime.datetime.now().strftime("%H:%M:%S")
    #broadcasting the emails
    for item in intersection_list: #the email just received
        if item not in excellist: #and it is not in the .csv logs (therefore it is not broadcasted yet)
            try:
                match = item
                tup = (match, timenow)
                os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
                writer = csv.writer(open(today + " Scheduling Broadcaster logs.csv","a", newline=""))
                writer.writerow(tup) #add the email subject in to .csv logs and save it
                country = match[0:2]
                messageupload = match
                print("\n" + match + " has been received")
            except Exception as e:
                messagebox.showwarning("Error", "Problem creating Scheduling Broadcaster logs in Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\nPlease enusre the directory where the tool will be stored exists\n\n"+ str(e))
                quit()
            
            webhook(country,messageupload) #broadcast to the country

        if item in excellist:
            continue
    
    print("\nSUCCESS!: All scheduling uploads received have been already broadcasted")
    











###################################
####Rostering Tracker##############
###################################

def rosteringtrackermain():
    #checking if we want to send Rostering Tracker webhooks after 3rd upload
    try:
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance')
        wMsgs = openpyxl.load_workbook("Scheduling_Messages.xlsx", data_only=True) #load workbook and telling Python to read only
        wMessages = wMsgs.worksheets[0]
        Having3rdupload  = wMessages.cell(row=28, column=2).value
    except Exception as e:
        messagebox.showwarning("Error", "File Scheduling_Messages.xlsx couldnt be found in \Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance in shared drive\n\n"+ str(e.__class__))
        quit()
    

    
    #creating file for logs in shared drive
    try:
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
        f = open(today + " Rostering Tracker Main logs.csv","a", newline="") 
        writer = csv.writer(f)
        f.close()
    except Exception as e:
        messagebox.showwarning("Error", "File Rostering Tracker Main logs.csv couldnt be created under Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\" + today + " Compliance Uploads logs \nMake sure such directory in share drive exists \n\n"+ str(e.__class__))
        quit()

    try:
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
        f = open(today + " Rostering Tracker__Do Not Open.csv","a", newline="") 
        writer = csv.writer(f)
        f.close()
    except Exception as e:
        messagebox.showwarning("Error", "File Rostering Tracker__Do Not Open.csv couldnt be created under Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\" + today + " Compliance Uploads logs \nMake sure such directory in share drive exists \n\n"+ str(e.__class__))
        quit()
    
    


    #get all values from csv
    #this csv will be updated adding new lines when rostering tracker has been ran for a country
    #it will help us finding if the rostering module was already executed for the country in question
    try:
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
        excellist = []
        logsfile = csv.reader(open(today + " Rostering Tracker Main logs.csv","r"))
        for row in logsfile:
            excellist.append(row[0]) #getting only the first column
        
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
        excellisttodisplay = []
        logsfile = csv.reader(open(today + " Rostering Tracker Main logs.csv","r"))
        for row in logsfile:
            excellisttodisplay.append(row) #getting all the row values
        
    except:
        messagebox.showwarning("Error", "File Rostering Tracker Main logs.csv couldnt be found under Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\" + today + " Compliance Uploads logs \nDelete all Rostering logs in such directory and restart tool \n\n"+ str(e.__class__))
        quit()

    
          
    try:
        #merging all lists of the potential email subjects that we can receive regarding scheduling uploads
        #if UK is having a 3rd upload, rostering webhooks will start to be sent once the 3rd upload is received
        uploadsubjectslist = []
        if Having3rdupload == "Yes":
            uploadsubjectslist = DS2nduploadlist + DS3rduploadlist
        else:
            uploadsubjectslist = DS2nduploadlist + DS2nduploadlistUK
        


        #tricky part: we are checking the emails subject received in the last 2 hrs, all the potential emails subject that can be received and the emails subject that have been already broadcasted and are stored in the logs file

        #seeing the common values between the emails subject just received and all the potential ones that can be received (that are stored in excel file)
        list_of_emails_set = set(list(dict.fromkeys(list_of_emails))) #emails received
        intersection = list_of_emails_set.intersection(uploadsubjectslist)
        intersection_list = list(intersection) #the common values are in this list
    
    except Exception as e:
        messagebox.showwarning("Error", "There is a variable which is not defined\nSeek BIA team assistance\n\n"+ str(e))
        quit()
    
            
    
    timenow = datetime.datetime.now().strftime("%H:%M:%S")
    weekday = datetime.datetime.today().weekday()
    
    #broadcasting rostering webhooks
    for item in intersection_list: #the email just received
        if item not in excellist: #and it is not in the .csv logs (therefore it is not broadcasted yet) 
            try:
                date = dateOFD
                if date == " ":
                    messagebox.showwarning("Error", "Note that the OFD couldn't be extracted from the scheduling email subject\nAsk Scheduling to resend email with the title correctly written\nAny space can stop the tool from reading the subject email correctly")
                    quit()
                match = item
                country = match[0:2]
                countries = []
                if country == "UK":
                    countries.extend(("UK","IE"))
                if country == "DE":
                    countries.extend(("DE","AT","NL"))
                    if weekday == 5:
                        print("\nIt's Saturday - removing NL from the list MEU. . .")
                        countries.remove("NL")
                        if date == datetime.datetime.strptime(today,"%Y-%m-%d") + timedelta(days=1):  #just in case it grabs DNL1 date on Saturday when both emails are sent at the same time
                            date = datetime.datetime.strptime(today,"%Y-%m-%d") + timedelta(days=2)
                if country == "FR":
                    countries.extend(("FR","BE"))
                if country == "IT":
                    countries.append("IT")
                if country == "ES":
                    countries.append("ES")
                tup = (match, timenow, countries, date)
            except Exception as e:
                messagebox.showwarning("Error", "Problem extracting the data to update excel logs\n\n"+ str(e))
                quit()
            
            try:
                #main version where the user can modify it
                os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
                with open(today + " Rostering Tracker Main logs.csv","a", newline="") as Rosteringlogs:
                    writer = csv.writer(Rosteringlogs)
                    writer.writerow(tup)
            except Exception as e:
                messagebox.showwarning("Error", "Problem creating Rostering Tracker Main logs.csv in Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\nPlease enusre the directory where the tool will be stored exists\n\n"+ str(e))
                quit() 
            
            #extended version where the broadcast times will be getting updated
            #this file will be used to get the last broadcasting times and see if those were done more than 30 mins ago. If so, we will send another broadcast and update the broadcasting time to the time now.
            try:
                tup = (match, str(timenow), countries, date)
                os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
                with open(today + " Rostering Tracker__Do Not Open.csv","a", newline="") as Rosteringlogsextended:
                    writer = csv.writer(Rosteringlogsextended)
                    writer.writerow(tup)
                    
                subjectmatch = match
                countrystr = ', '.join(map(str, countries))
                print("\nRostering tracker required for " + str(countrystr) + " ...")
            
            except Exception as e:
                messagebox.showwarning("Error", "Problem creating Rostering Tracker__Do Not Open.csv in Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\nPlease enusre the directory where the tool will be stored exists\n\n"+ str(e))
                quit()

            AutomatedRosteringTracker(countries,date,subjectmatch,today) #broadcast to the country
        
        if item in excellist:
            continue
    
    #for NL on Saturday
    DESaturday = "DE - Next Day - Daily 2nd run "
    NLSaturday = "NL - Next Day - Daily 2nd run "
    
    if weekday == 5 and DESaturday in excellist and NLSaturday not in excellist:
        print("\nIt's Saturday - removing NL from the list MEU. . .")
        countries = ["NL"]
        try:
            date = datetime.datetime.strptime(today,"%Y-%m-%d") + timedelta(days=1)
            tup = (NLSaturday, timenow, countries, date)
            countrystr = ', '.join(map(str, countries))
            subjectmatch = NLSaturday
            print("\nRostering tracker required for " + str(countrystr) + " ...")
        except Exception as e:
            messagebox.showwarning("Error", "Problem extracting the data to update excel logs\n\n"+ str(e))
            quit()
        
        try:
            #main version where the user can modify it
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
            with open(today + " Rostering Tracker Main logs.csv","a", newline="") as Rosteringlogs:
                writer = csv.writer(Rosteringlogs)
                writer.writerow(tup)
        except Exception as e:
            messagebox.showwarning("Error", "Problem creating Rostering Tracker Main logs.csv in Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\nPlease enusre the directory where the tool will be stored exists\n\n"+ str(e))
            quit()

        try:
            #extended version where the broadcast times will be getting updated
            tup = (NLSaturday, str(timenow), countries, date)
            os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs')
            with open(today + " Rostering Tracker__Do Not Open.csv","a", newline="") as Rosteringlogsextended:
                writer = csv.writer(Rosteringlogsextended)
                writer.writerow(tup)
        except Exception as e:
            messagebox.showwarning("Error", "Problem creating Rostering Tracker__Do Not Open.csv in Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\nPlease enusre the directory where the tool will be stored exists\n\n"+ str(e))
            quit()
        
        AutomatedRosteringTracker(countries,date,subjectmatch,today) #broadcast to the country
    
    


    #every 30 mins rostering tracker will be sent
    timenow = datetime.datetime.now().strftime("%H:%M:%S")
    try:
        logsfile = csv.reader(open(today + " Rostering Tracker__Do Not Open.csv","r"))
    except Exception as e:
        messagebox.showwarning("Error", "Problem creating Rostering Tracker__Do Not Open.csv in Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\nPlease enusre the directory where the tool will be stored exists\n\n"+ str(e))
        quit()

    for row in logsfile:
        for countrysearch in excellist:
            if row[0] == countrysearch:
                timenowroster = datetime.datetime.strptime(timenow,"%H:%M:%S")
                try:
                    timenextrostering = datetime.datetime.strptime(row[1][11:19],"%H:%M:%S")  + timedelta(minutes = 29) 
                except:
                    timenextrostering = datetime.datetime.strptime(timenow,"%H:%M:%S") - timedelta(minutes = 10) 
                
                try:
                    timenowroster = timenowroster.time()
                    timenextrostering = timenextrostering.time()

                    Rosteringinfo = row
                    countries = Rosteringinfo[2] #column C
                    date1 = Rosteringinfo[3] #column D
                    subjectmatch = Rosteringinfo[0]
                
                except Exception as e:
                    messagebox.showwarning("Error", "There is a variable which is not defined\nSeek BIA team assistance\n\n"+ str(e))
                    quit()

        
                # if the last rostering webhook sent was 29 mins ago, trigger Rostering Tracker
                if timenowroster > timenextrostering:
                    #extracting OFD date before triggering Rostering Tracker
                    try:
                        datecomplete= date1[0:10]
                    except Exception as e:
                        messagebox.showwarning("Error", "Problem reading OFD in Rostering Tracker__Do Not Open.csv in Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\nPlease delete the existent Rostering Tracker__Do Not Open.csv and restart app\n\n"+ str(e))
                        quit()
                    
                    try:
                        date = datetime.datetime.strptime(datecomplete,"%Y-%m-%d")
                    except:
                        date = datetime.datetime.strptime(datecomplete,"%d/%m/%Y")
                    
                    
                    print("\nRostering tracker required for " + str(countries) + " ...")
                    AutomatedRosteringTracker(countries,date,subjectmatch,today)







###################################
#SA Planning Part by Adrian Stoica#
###################################

def SAPlanningAlert():
    SAPlanningAlertAutomated(today,currentdirectory)

    

    

def main():
    while True:
        
        SchedulingBroadcaster()
        rosteringtrackermain()  
        countdown(int(Searchsecs))
        
        SchedulingBroadcaster()
        rosteringtrackermain()
        SAPlanningAlert()   
        SWATmain() #SWAT metrics module
        countdown(int(Searchsecs)) 
        
        
        


if __name__ == "__main__":
    main()
    


